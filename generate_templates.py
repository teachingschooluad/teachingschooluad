"""Generate Excel import templates for ePKL system."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), 'public', 'templates')
os.makedirs(TEMPLATES_DIR, exist_ok=True)

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4648D4', end_color='4648D4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
EXAMPLE_FONT = Font(name='Arial', italic=True, size=10, color='888888')

def make_template(filename, headers, examples, col_widths):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    for col_idx, example in enumerate(examples, 1):
        cell = ws.cell(row=2, column=col_idx, value=example)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER
    
    filepath = os.path.join(TEMPLATES_DIR, filename)
    wb.save(filepath)
    print(f'  [OK] Created {filepath}')

print('Generating ePKL Excel templates...')

make_template('template_siswa.xlsx',
    headers=['nama', 'username', 'nisn', 'class_id', 'company_id', 'guru_id'],
    examples=['Budi Santoso', 'budi123', '1234567890', '1', '2', '3'],
    col_widths=[25, 18, 18, 12, 14, 12])

make_template('template_guru.xlsx',
    headers=['nama', 'username', 'nip'],
    examples=['Pak Ahmad', 'ahmad_guru', '198501012010011234'],
    col_widths=[25, 18, 25])

make_template('template_perusahaan.xlsx',
    headers=['nama', 'alamat', 'pemilik', 'username', 'gps1', 'gps2', 'gps3', 'gps4'],
    examples=['PT Maju Jaya', 'Jl. Merdeka No. 1, Cilacap', 'Ir. Suharto', 'majujaya', '-7.728, 109.01', '', '', ''],
    col_widths=[25, 35, 20, 15, 18, 18, 18, 18])

print('Done!')
